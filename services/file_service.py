import json
import os
import openpyxl
import configparser
import re
import pandas as pd


class FileService:
    """
    统一的文件操作服务类，负责 JSON, Excel, INI 等文件的读取和保存。
    """
    APP_NAME = "oahutree_spiderling"

    @staticmethod
    def get_app_home():
        """获取用户目录下的应用主目录"""
        return os.path.expanduser(os.path.join("~", FileService.APP_NAME))

    @staticmethod
    def get_config_dir():
        """获取配置目录"""
        return os.path.join(FileService.get_app_home(), "config")

    @staticmethod
    def initialize_app_data():
        """
        初始化应用数据。如果系统主目录下没有对应目录，则创建并从包内资源拷贝初始配置。
        """
        import shutil
        app_home = FileService.get_app_home()
        config_dir = FileService.get_config_dir()
        locales_dir = os.path.join(config_dir, "locales")
        cache_dir = os.path.join(app_home, ".cache")
        db_dir = os.path.join(app_home, ".sqlitedb")

        # 创建目录结构
        for d in [app_home, config_dir, locales_dir, cache_dir, db_dir]:
            os.makedirs(d, exist_ok=True)

        # 需要拷贝的基础文件（相对于当前运行目录）
        # 只有这些文件需要拷贝到用户目录并支持重置
        base_files = [
            "settings.ini", "languages.json", "browser_config.json", "stage_type.json"
        ]
        for f in base_files:
            src = os.path.join("config", f)
            dst = os.path.join(config_dir, f)
            if os.path.exists(src) and not os.path.exists(dst):
                shutil.copy2(src, dst)

        # 拷贝 locales
        src_locales = os.path.join("config", "locales")
        if os.path.exists(src_locales):
            for f in os.listdir(src_locales):
                if f.endswith(".json"):
                    src_f = os.path.join(src_locales, f)
                    dst_f = os.path.join(locales_dir, f)
                    if not os.path.exists(dst_f):
                        shutil.copy2(src_f, dst_f)

    @staticmethod
    def get_resetable_files():
        """获取允许重置的文件列表"""
        return [
            "settings.ini", "languages.json", "browser_config.json", "stage_type.json"
        ]

    @staticmethod
    def reset_file(file_name):
        """
        重置单个配置文件。
        """
        import shutil
        src = os.path.join("config", file_name)
        dst = os.path.join(FileService.get_config_dir(), file_name)
        if os.path.exists(src):
            shutil.copy2(src, dst)
            return True
        return False

    @staticmethod
    def reset_locales():
        """
        重置 locales 目录下的所有语言文件。
        """
        import shutil
        src_locales = os.path.join("config", "locales")
        dst_locales = os.path.join(FileService.get_config_dir(), "locales")
        if os.path.exists(src_locales):
            os.makedirs(dst_locales, exist_ok=True)
            for f in os.listdir(src_locales):
                if f.endswith(".json"):
                    shutil.copy2(os.path.join(src_locales, f), os.path.join(dst_locales, f))
            return True
        return False

    @staticmethod
    def get_config_path(file_name):
        """
        获取配置文件的完整路径。
        部分核心配置文件从程序根目录的 config 加载（全局共享，不可重置）。
        用户个性化配置从用户目录下的 config 加载（可重置）。
        """
        global_configs = [
            "menu.json", "fields.json", "actions.json", "ignore_type.json",
            "source_type.json", "stage_type.json"
        ]
        if file_name in global_configs:
            return os.path.join("config", file_name)
        return os.path.join(FileService.get_config_dir(), file_name)

    @staticmethod
    def get_last_workbench_file():
        """从 settings.ini 获取上次打开的工作台文件"""
        settings_path = FileService.get_config_path("settings.ini")
        settings = FileService.load_ini(settings_path)
        path = settings.get("General", {}).get("last_workbench_file", "")
        if path and os.path.exists(path):
            return path
        return None

    @staticmethod
    def set_last_workbench_file(file_path):
        """保存当前工作台文件路径到 settings.ini"""
        settings_path = FileService.get_config_path("settings.ini")
        settings = FileService.load_ini(settings_path)
        if "General" not in settings:
            settings["General"] = {}
        settings["General"]["last_workbench_file"] = file_path
        FileService.save_ini(settings_path, settings)

    @staticmethod
    def load_json(file_path, default_data=None):
        """
        加载 JSON 文件。

        Args:
            file_path (str): 文件路径。
            default_data: 文件不存在时返回的默认数据。

        Returns:
            dict/list: 解析后的数据。
        """
        if not os.path.exists(file_path):
            return default_data if default_data is not None else {}
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading JSON from {file_path}: {e}")
            return default_data if default_data is not None else {}

    @staticmethod
    def save_json(file_path, data):
        """
        保存数据到 JSON 文件。

        Args:
            file_path (str): 文件路径。
            data: 要保存的数据。
        """
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving JSON to {file_path}: {e}")

    @staticmethod
    def load_excel(file_path, fields_config=None):
        """
        加载 Excel 文件并根据字段配置映射数据。

        Args:
            file_path (str): 文件路径。
            fields_config (list): 可选的字段配置，用于映射数据顺序。

        Returns:
            dict: 键为 Sheet 名称，值为行数据列表的字典。
        """
        if not os.path.exists(file_path):
            return {}

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            data = {}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.max_row < 1:
                    continue

                # 获取表头
                excel_headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]

                if fields_config:
                    # 建立映射：key -> excel_column_index
                    col_map = {}
                    for col_idx, header in enumerate(excel_headers):
                        match = re.search(r"\(([^)]+)\)$", header)
                        if match:
                            key = match.group(1)
                            col_map[key] = col_idx
                        else:
                            for f in fields_config:
                                if f.get("key") == header:
                                    col_map[header] = col_idx
                                    break

                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not any(row): continue
                        mapped_row = []
                        for f in fields_config:
                            key = f.get("key")
                            val = ""
                            if key in col_map:
                                idx = col_map[key]
                                if idx < len(row):
                                    val = row[idx]
                            mapped_row.append(val)
                        rows.append(mapped_row)
                    data[sheet_name] = rows
                else:
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            rows.append(list(row))
                    data[sheet_name] = rows

            wb.close()
            return data
        except Exception as e:
            print(f"Error loading Excel from {file_path}: {e}")
            return {}

    @staticmethod
    def save_excel(file_path, data, fields, t=None):
        """
        将数据保存到 Excel 文件。

        Args:
            file_path (str): 文件路径。
            data (dict): Sheet 数据。
            fields (list): 字段配置。
            t: 翻译函数。
        """
        try:
            wb = openpyxl.Workbook()
            if wb.active:
                wb.remove(wb.active)

            headers = []
            for f in fields:
                key = f.get("key")
                label = key
                if t:
                    i18n_key = f.get("i18n_key")
                    if i18n_key:
                        label = t(i18n_key)
                headers.append(f"{label} ({key})")

            for sheet_name, rows in data.items():
                sheet = wb.create_sheet(title=sheet_name)
                sheet.append(headers)
                for i, row in enumerate(rows):
                    if len(row) > 0:
                        row[0] = i + 1
                    sheet.append(row)

            wb.save(file_path)
            wb.close()
        except Exception as e:
            print(f"Error saving Excel to {file_path}: {e}")

    @staticmethod
    def load_ini(file_path):
        """
        加载 INI 配置文件。
        """
        if not os.path.exists(file_path):
            return {}
        try:
            config = configparser.ConfigParser()
            config.read(file_path, encoding="utf-8")
            return {s: dict(config.items(s)) for s in config.sections()}
        except Exception as e:
            print(f"Error loading INI from {file_path}: {e}")
            return {}

    @staticmethod
    def save_ini(file_path, data):
        """
        保存数据到 INI 配置文件。
        """
        try:
            config = configparser.ConfigParser()
            for section, options in data.items():
                config[section] = options
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, "w", encoding="utf-8") as f:
                config.write(f)
        except Exception as e:
            print(f"Error saving INI to {file_path}: {e}")

    @staticmethod
    def ensure_excel_template(file_path, fields, t=None):
        """
        确保 Excel 模板文件存在。如果不存在，则根据字段配置创建一个。
        """
        if os.path.exists(file_path):
            return

        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "ScrapingSteps"

            headers = []
            for f in fields:
                key = f.get("key")
                label = key
                if t:
                    i18n_key = f.get("i18n_key")
                    if i18n_key:
                        label = t(i18n_key)
                headers.append(f"{label} ({key})")

            sheet.append(headers)
            wb.save(file_path)
            wb.close()
        except Exception as e:
            print(f"Error creating Excel template at {file_path}: {e}")

    @staticmethod
    def del_cache(config_path, file_name):
        """
        删除指定的缓存文件。
        """
        try:
            full_path = os.path.join(config_path, file_name)
            if os.path.exists(full_path):
                os.remove(full_path)
                return True
            return False
        except Exception as e:
            print(f"Error deleting cache file {file_name}: {e}")
            return False

    @staticmethod
    def write_cache(config_path, file_name, data, delete_existing=True):
        """
        将数据写入缓存文本文件。
        支持 string, int, list 或 dataframe。
        使用 CSV 格式处理含有逗号的情况。
        delete_existing: 如果为 True 则覆盖，为 False 则进行合并/相加。
        """
        try:
            full_path = os.path.join(config_path, file_name)

            # 如果不删除且文件存在，则先尝试读取并合并数据
            if not delete_existing and os.path.exists(full_path):
                # 根据输入 data 的类型决定如何读取旧数据
                d_type = "string"
                is_df = False
                try:

                    if isinstance(data, pd.DataFrame):
                        d_type = "dataframe"
                        is_df = True
                    elif isinstance(data, int):
                        d_type = "int"
                    elif isinstance(data, (list, tuple)):
                        d_type = "list"
                except:
                    if isinstance(data, int):
                        d_type = "int"
                    elif isinstance(data, (list, tuple)):
                        d_type = "list"

                old_data = FileService.read_cache(config_path, file_name, d_type)

                if old_data is not None:
                    if d_type == "int":
                        data = old_data + data
                    elif d_type == "list":
                        data = list(old_data) + list(data)
                    elif is_df:
                        # 如果旧数据和新数据都只有一行，则进行列合并（适用于分步构建单条记录的场景）
                        if len(old_data) == 1 and len(data) == 1:
                            try:
                                # 1. 强制索引对齐，确保 combine_first 能按位置合并
                                data.index = old_data.index

                                # 2. 将空字符串统一转为 NaN，否则 combine_first 不会进行覆盖合并
                                import numpy as np
                                data = data.replace(r'^\s*$', np.nan, regex=True)
                                old_data = old_data.replace(r'^\s*$', np.nan, regex=True)

                                # 3. 使用 combine_first 合并数据
                                data = data.combine_first(old_data)
                            except Exception as merge_err:
                                print(f"Merge error, falling back to concat: {merge_err}")
                                data = pd.concat([old_data, data], ignore_index=True)
                        else:
                            data = pd.concat([old_data, data], ignore_index=True)
                    else:
                        data = str(old_data) + "\n" + str(data)

            os.makedirs(config_path, exist_ok=True)
            # content = ""
            # 处理 DataFrame (尝试导入 pandas)
            try:

                if isinstance(data, pd.DataFrame):
                    # 使用 to_csv 更好地处理逗号和引号
                    content = data.to_csv(index=False)
                else:
                    raise ImportError
            except (ImportError, NameError):
                if isinstance(data, (list, tuple)):
                    # 如果列表中的项含有逗号，只要作为整行写入，读取时按行切分即可
                    content = "\n".join(str(item) for item in data)
                else:
                    content = str(data)

            with open(full_path, "w", encoding="utf-8") as f:
                f.write(content)
            return True
        except Exception as e:
            print(f"Error writing cache to {file_name}: {e}")
            return False

    @staticmethod
    def read_cache(config_path, file_name, data_type="string", record_index=None):
        """
        读取缓存文件并转换为指定类型。
        支持 'string', 'int', 'list', 'dataframe'。
        record_index: 获取指定的某一条记录（索引从0开始）。
        """
        try:
            full_path = os.path.join(config_path, file_name)
            if not os.path.exists(full_path):
                return None

            with open(full_path, "r", encoding="utf-8") as f:
                content = f.read()

            result = None
            if data_type == "int":
                try:
                    result = int(content.strip())
                except:
                    result = 0
            elif data_type == "list":
                result = [line for line in content.splitlines() if line]
            elif data_type == "dataframe":
                try:
                    import pandas as pd
                    from io import StringIO
                    # 如果内容看起来像 CSV（有列头且 write_cache 使用了 to_csv）
                    result = pd.read_csv(StringIO(content))
                except Exception as e:
                    print(f"Error converting to DataFrame: {e}")
                    result = None
            else:
                result = content

            # 处理特定记录提取
            if record_index is not None and result is not None:
                if data_type == "int":
                    return result  # int 忽略 record_index

                try:
                    if data_type == "list":
                        return result[record_index] if 0 <= record_index < len(result) else None
                    elif data_type == "dataframe":
                        return result.iloc[record_index] if 0 <= record_index < len(result) else None
                    elif data_type == "string":
                        lines = result.splitlines()
                        return lines[record_index] if 0 <= record_index < len(lines) else None
                except Exception as e:
                    print(f"Error fetching record at index {record_index}: {e}")
                    return None

            return result
        except Exception as e:
            print(f"Error reading cache from {file_name}: {e}")
            return None

    @staticmethod
    def get_cache_count(config_path, file_name, data_type="string"):
        """
        获取缓存文件中的记录总数（行数）。
        支持 string, int, list, dataframe。
        """
        data = FileService.read_cache(config_path, file_name, data_type)
        if data is None:
            return 0

        if data_type == "int":
            return 1
        elif data_type == "list":
            return len(data)
        elif data_type == "dataframe":
            try:
                return len(data)
            except:
                return 0
        elif data_type == "string":
            return len(data.splitlines())
        return 0

    @staticmethod
    def get_browser_user_data_dir():
        """
        读取 browser_config.json 文件中定义的 user_data_dir 值。
        """
        config_file = FileService.get_config_path("browser_config.json")
        config = FileService.load_json(config_file)
        # 默认返回家目录应用文件夹下的 chrome_profile
        user_data_dir = config.get("user_data_dir", "chrome_profile")
        if not os.path.isabs(user_data_dir):
            user_data_dir = os.path.join(FileService.get_app_home(), user_data_dir)
        return user_data_dir

    @staticmethod
    def get_db_conn_string():
        """
        自动读取数据库配置文件并生成当前选中的连接字符串。
        """
        config_file = FileService.get_config_path("db_config.json")

        config = FileService.load_json(config_file)
        if not config or "databases" not in config:
            return ""

        idx = config.get("current_index", 0)
        databases = config.get("databases", [])

        if idx < 0 or idx >= len(databases):
            return ""

        db_cfg = databases[idx]
        db_type = db_cfg.get("type")
        host = db_cfg.get("host")
        port = db_cfg.get("port")
        db_name = db_cfg.get("database")
        user = db_cfg.get("username")
        pwd = db_cfg.get("password")
        params = db_cfg.get("params", "")

        return FileService.generate_conn_string(db_type, host, port, db_name, user, pwd, params)

    @staticmethod
    def generate_conn_string(db_type, host, port, db_name, user, pwd, params=""):
        """
        根据参数生成 SQLAlchemy 格式的连接字符串。
        """
        conn_str = ""
        if db_type == "SQLite":
            # 如果是相对路径，则相对于 APP_HOME
            if not os.path.isabs(os.path.expanduser(db_name)):
                db_name = os.path.join(FileService.get_app_home(), db_name)
            full_path = os.path.expanduser(db_name)
            conn_str = f"sqlite:///{full_path}"
        elif db_type == "MySQL":
            conn_str = f"mysql+pymysql://{user}:{pwd}@{host}:{port}/{db_name}"
        elif db_type == "PostgreSQL":
            conn_str = f"postgresql://{user}:{pwd}@{host}:{port}/{db_name}"
        elif db_type == "SQL Server":
            conn_str = f"mssql+pymssql://{user}:{pwd}@{host}:{port}/{db_name}"
        elif db_type == "Oracle":
            conn_str = f"oracle+cx_oracle://{user}:{pwd}@{host}:{port}/?service_name={db_name}"

        if params and conn_str:
            if "?" in conn_str:
                conn_str += f"&{params}"
            else:
                conn_str += f"?{params}"
        return conn_str
